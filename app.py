# ================= USER CONFIG =================
KEY_COLUMN = "P"
DATE_COLUMN = "G"
NUMBER_COLUMNS = ("J", "P")
DELETE_TOP_ROWS = 17
FREEZE_PANE = "A4"

COLUMN_WIDTH = 15
ROW_HEIGHT = 15
DATE_FORMAT = "MM/DD/YYYY"
NUMBER_FORMAT = "#,##0.00"
# ==============================================

import streamlit as st
import os
import calendar
from datetime import datetime
from openpyxl import load_workbook, Workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Border, Side
import io


def clean_sheet(sheet):
    thin_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )

    for m in list(sheet.merged_cells):
        sheet.unmerge_cells(str(m))

    sheet.delete_rows(1, DELETE_TOP_ROWS)

    for r in range(sheet.max_row, 0, -1):
        if sheet.cell(r, 1).value in (None, ""):
            sheet.delete_rows(r)

    key_idx = ord(KEY_COLUMN) - ord("A") + 1
    last_row = 0
    for r in range(sheet.max_row, 0, -1):
        if sheet.cell(r, key_idx).value not in (None, ""):
            last_row = r
            break
    if last_row and last_row < sheet.max_row:
        sheet.delete_rows(last_row + 1, sheet.max_row - last_row)

    for c in range(sheet.max_column, 0, -1):
        if all(sheet.cell(r, c).value in (None, "") for r in range(1, sheet.max_row + 1)):
            sheet.delete_cols(c)

    for r in range(sheet.max_row, 0, -1):
        if all(sheet.cell(r, c).value in (None, "") for c in range(1, sheet.max_column + 1)):
            sheet.delete_rows(r)

    for col in sheet.columns:
        sheet.column_dimensions[get_column_letter(col[0].column)].width = COLUMN_WIDTH
    for rdim in sheet.row_dimensions.values():
        rdim.height = ROW_HEIGHT

    sheet._images.clear()

    for row in sheet.iter_rows():
        for cell in row:
            cell.border = thin_border


def map_filename(name):
    NAME_MAP = {
        "DIVRE": "Kantor Divre Jawa Timur",
        "BWB": "KPH Banyuwangi Barat",
        "BWS": "KPH Banyuwangi Selatan",
        "BWU": "KPH Banyuwangi Utara",
        "BTR": "KPH Blitar",
        "BNG": "KPH Bojonegoro",
        "BDO": "KPH Bondowoso",
        "JTR": "KPH Jatirogo",
        "JBR": "KPH Jember",
        "JBG": "KPH Jombang",
        "KDR": "KPH Kediri",
        "LWU": "KPH Lawu",
        "MDN": "KPH Madiun",
        "MDR": "KPH Madura",
        "MLG": "KPH Malang",
        "MJK": "KPH Mojokerto",
        "NGK": "KPH Nganjuk",
        "NGW": "KPH Ngawi",
        "PBO": "KPH Probolinggo",
        "PDG": "KPH Padangan",
        "PRG": "KPH Parengan",
        "PSU": "KPH Pasuruan",
        "SRD": "KPH Saradan",
        "TBN": "KPH Tuban",
        "DEPREN": "Departemen Perencanaan Jawa Timur",
    }
    return NAME_MAP.get(name, name)


def process_files(uploaded_files):
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "COMBINED"
    out_row = 1

    for uploaded_file in uploaded_files:
        src_wb = load_workbook(uploaded_file)

        for src_ws in src_wb.worksheets:
            clean_sheet(src_ws)
            max_col = src_ws.max_column
            base = os.path.splitext(uploaded_file.name)[0]

            for r in range(1, src_ws.max_row + 1):
                out_ws.cell(out_row, 1, src_ws.cell(r, 1).value)
                out_ws.cell(out_row, 2, map_filename(base))

                for c in range(2, max_col + 1):
                    out_ws.cell(out_row, c + 1, src_ws.cell(r, c).value)

                out_row += 1
    
    for m in list(out_ws.merged_cells):
    out_ws.unmerge_cells(str(m))

    out_ws.insert_rows(1, 3)

    start = ord(NUMBER_COLUMNS[0]) - ord("A") + 1
    end = ord(NUMBER_COLUMNS[1]) - ord("A") + 1
    for r in range(4, out_ws.max_row + 1):
        for c in range(start, end + 1):
            cell = out_ws.cell(r, c)
            if isinstance(cell.value, (int, float)):
                cell.number_format = NUMBER_FORMAT

    date_idx = ord(DATE_COLUMN) - ord("A") + 1
    for r in range(4, out_ws.max_row + 1):
        cell = out_ws.cell(r, date_idx)
        v = cell.value
        if isinstance(v, str):
            try:
                d = datetime.strptime(v.strip(), "%b-%y")
                cell.value = datetime(d.year, d.month, calendar.monthrange(d.year, d.month)[1])
                cell.number_format = DATE_FORMAT
            except:
                pass
        elif isinstance(v, datetime):
            cell.number_format = DATE_FORMAT

    headers = [
        "No. Urut", "Satker", "Golongan Penyusutan", "Nomor Rekening",
        "Jenis Aktiva Tetap", "Nomor Aktiva Tetap", "Tahun Perolehan",
        "Masa Manfaat", "Tarif Penyusutan", "Nilai Perolehan",
        "Nilai Buku s/d Bulan Lalu",
        "Penyusutan s/d Bulan Lalu", "Biaya Peny. dalam Bulan",
        "Biaya Peny. s/d Bulan", "Akumulasi Penyusutan s/d Bulan",
        "Nilai Buku s/d Bulan ini", "Nomor Rekening Penyusutan"
    ]

    for i, h in enumerate(headers, start=1):
        out_ws.cell(1, i, h)

    counter = 1
    for r in range(4, out_ws.max_row + 1):
        out_ws.cell(r, 1, counter)
        counter += 1

    out_ws.freeze_panes = out_ws[FREEZE_PANE]

    output = io.BytesIO()
    out_wb.save(output)
    return output.getvalue()


# ================= STREAMLIT UI =================
st.title("Excel Combiner & Cleaner")

uploaded_files = st.file_uploader(
    "Upload Excel files",
    type=["xlsx"],
    accept_multiple_files=True
)

if uploaded_files:
    result = process_files(uploaded_files)
    st.download_button(
        "Download Combined Excel",
        data=result,
        file_name="FINAL_CLEAN.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
